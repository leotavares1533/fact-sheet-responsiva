#!/usr/bin/env python
"""Import Confina funding schedules from the user workbook into the local UI data file."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SAO_PAULO_TIMEZONE = dt.timezone(dt.timedelta(hours=-3), "America/Sao_Paulo")
EXPORT_VARIABLE = "ceresFundingData"

GROUPS = [
    {
        "id": "confina-cra-65-200",
        "investor": "Ceres",
        "name": "Confina CRA 65a - R$ 200MM",
        "shortName": "CRA 65a R$ 200MM",
        "account": "A cadastrar",
        "sheets": ["CRA CERES 65 200mm"],
    },
    {
        "id": "confina-cra-65-80",
        "investor": "Ceres",
        "name": "Confina CRA 65a - R$ 80MM",
        "shortName": "CRA 65a R$ 80MM",
        "account": "A cadastrar",
        "sheets": ["CRA CERES 65 80mm"],
    },
    {
        "id": "confina-cra-42-50",
        "investor": "Ceres",
        "name": "Confina CRA 42a - R$ 50MM",
        "shortName": "CRA 42a R$ 50MM",
        "account": "A cadastrar",
        "sheets": ["CPRF CONFINA CRA 42"],
    },
    {
        "id": "confina-cras-carteira-100",
        "investor": "Ceres",
        "name": "Confina CRAs Carteira - R$ 100MM",
        "shortName": "CRAs Carteira R$ 100MM",
        "account": "A cadastrar",
        "sheets": [
            "CRA Carteira 100mm 1 (1.8%)",
            "CRA Carteira 100mm 2 (1.8%)",
            "CRA Carteira 100mm 3 (1.8%)",
            "CRA Carteira 100mm 4 (1.8%)",
            "CRA Carteira 100mm 5 (1.8%)",
            "CRA Carteira 100mm 6 (1.8%)",
        ],
    },
    {
        "id": "confina-cras-carteira-50",
        "investor": "Ceres",
        "name": "Confina CRAs Carteira - R$ 50MM",
        "shortName": "CRAs Carteira R$ 50MM",
        "account": "A cadastrar",
        "sheets": [
            "CRA Carteira 50mm 1 (1.8%)",
            "CRA Carteira 50mm 2 (1.8%)",
        ],
    },
    {
        "id": "confina-cprf-100",
        "investor": "BTG",
        "name": "Confina CPRF - R$ 100MM",
        "shortName": "CPRF R$ 100MM",
        "account": "A cadastrar",
        "sheets": ["CPRF 100MM BTG"],
    },
    {
        "id": "confina-cprf-50",
        "investor": "BTG",
        "name": "Confina CPRF - R$ 50MM",
        "shortName": "CPRF R$ 50MM",
        "account": "A cadastrar",
        "sheets": ["CPRF 50MM BTG"],
    },
]


def normalize_label(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def as_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    if isinstance(value, str):
        text = value.replace("R$", "").replace("%", "").replace(" ", "")
        if "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def clean_money(value: Any) -> float:
    number = as_float(value)
    if number is None or abs(number) < 1:
        return 0.0
    return float(number)


def iso_date(value: dt.date | None) -> str | None:
    return value.isoformat() if value else None


def find_labeled_value(sheet: Any, label: str) -> Any:
    target = normalize_label(label)
    for row in range(1, min(sheet.max_row, 12) + 1):
        for column in range(1, min(sheet.max_column, 8) + 1):
            if normalize_label(sheet.cell(row, column).value) == target:
                return sheet.cell(row, column + 1).value
    return None


def find_header_row(sheet: Any) -> tuple[int, dict[str, int]]:
    for row in range(1, min(sheet.max_row, 20) + 1):
        headers = {}
        for column in range(1, sheet.max_column + 1):
            label = normalize_label(sheet.cell(row, column).value)
            if label:
                headers[label] = column
        if "dia" in headers and "saldo devedor" in headers:
            return row, headers
    raise ValueError(f"Nao encontrei cabecalho de cronograma na aba {sheet.title}")


def column_for(headers: dict[str, int], *labels: str) -> int | None:
    for label in labels:
        normalized = normalize_label(label)
        if normalized in headers:
            return headers[normalized]
    return None


def funding_rate_terms(sheet: Any) -> dict[str, Any]:
    cdi = as_float(find_labeled_value(sheet, "CDI"))
    spread = as_float(find_labeled_value(sheet, "Taxa"))
    nominal = as_float(find_labeled_value(sheet, "Taxa nominal"))
    if cdi is not None and nominal is not None:
        monthly = (math.pow(1 + nominal, 1 / 12) - 1) * 100
        return {
            "fundingRate": monthly,
            "fundingRateLabel": f"CDI {cdi * 100:.2f}% a.a. + spread {(spread or 0) * 100:.2f}% a.a.",
            "fundingRateType": "cdi_spread",
            "cdiRate": cdi,
            "spreadRate": spread or 0.0,
            "annualRate": nominal,
        }
    if spread is None:
        return {
            "fundingRate": 0.0,
            "fundingRateLabel": "Taxa a cadastrar",
            "fundingRateType": "unknown",
            "cdiRate": None,
            "spreadRate": None,
            "annualRate": None,
        }
    monthly = spread * 100
    annual = math.pow(1 + spread, 12) - 1
    return {
        "fundingRate": monthly,
        "fundingRateLabel": f"{monthly:.2f}% a.m.",
        "fundingRateType": "monthly_fixed",
        "cdiRate": None,
        "spreadRate": None,
        "annualRate": annual,
    }


def sheet_position(sheet: Any, position_date: dt.date) -> dict[str, Any]:
    header_row, headers = find_header_row(sheet)
    dia_col = headers["dia"]
    saldo_col = headers["saldo devedor"]
    juros_col = column_for(headers, "Juros Dia", "Juros")
    juros_pago_col = column_for(headers, "Juros Pago")
    amort_col = column_for(headers, "Amortizacao", "Amortizacao ", "Amortização")
    parcela_col = column_for(headers, "Parcela")
    factor_col = column_for(headers, "Fator diario")
    rate_col = column_for(headers, "Taxa")
    spread_col = column_for(headers, "Spread")

    workbook_history: list[dict[str, Any]] = []
    events: list[dict[str, Any]] = []
    rate_schedule: list[dict[str, Any]] = []
    interest_paid = 0.0
    amortization_paid = 0.0
    installment_paid = 0.0
    final_payment_date: dt.date | None = None
    start_date: dt.date | None = None
    accrue_on_start_date = False
    rate_terms = funding_rate_terms(sheet)
    base_days = int(as_float(find_labeled_value(sheet, "Base")) or 360)

    previous_balance: float | None = None
    for row in range(header_row + 1, sheet.max_row + 1):
        current_date = as_date(sheet.cell(row, dia_col).value)
        if not current_date:
            continue
        if start_date is None:
            start_date = current_date
        balance = clean_money(sheet.cell(row, saldo_col).value)
        interest = clean_money(sheet.cell(row, juros_col).value) if juros_col else 0.0
        interest_paid_row = clean_money(sheet.cell(row, juros_pago_col).value) if juros_pago_col else 0.0
        amortization = clean_money(sheet.cell(row, amort_col).value) if amort_col else 0.0
        installment = clean_money(sheet.cell(row, parcela_col).value) if parcela_col else 0.0
        daily_factor = as_float(sheet.cell(row, factor_col).value) if factor_col else None
        cdi_rate = as_float(sheet.cell(row, rate_col).value) if rate_col else None
        spread_rate = as_float(sheet.cell(row, spread_col).value) if spread_col else None
        if current_date == start_date and interest > 0.01:
            accrue_on_start_date = True

        if current_date <= position_date:
            workbook_history.append({"date": current_date.isoformat(), "balance": round(balance, 2)})
            if rate_terms["fundingRateType"] == "cdi_spread" and daily_factor is not None:
                rate_schedule.append(
                    {
                        "date": current_date.isoformat(),
                        "dailyRate": round(float(daily_factor), 14),
                        "cdiRate": cdi_rate,
                        "spreadRate": spread_rate,
                    }
                )
            if installment > 0.01 or amortization > 0.01 or interest_paid_row > 0.01:
                if juros_pago_col is None and installment > amortization:
                    interest_paid_row = installment - amortization
                amount = installment if installment > 0.01 else interest_paid_row + amortization
                events.append(
                    {
                        "date": current_date.isoformat(),
                        "amount": round(amount, 2),
                        "interestPaid": round(interest_paid_row, 2),
                        "amortization": round(amortization, 2),
                        "sourceRow": row,
                    }
                )
            interest_paid += interest_paid_row
            amortization_paid += amortization
            installment_paid += installment
        elif final_payment_date is None and previous_balance and previous_balance > 1 and balance <= 1:
            final_payment_date = current_date

        previous_balance = balance

    if not workbook_history or start_date is None:
        raise ValueError(f"Nao encontrei posicao ate {position_date.isoformat()} na aba {sheet.title}")

    principal = clean_money(find_labeled_value(sheet, "Valor aquisicao"))
    issue_date = as_date(find_labeled_value(sheet, "Emissao"))
    maturity_date = as_date(find_labeled_value(sheet, "Vencimento")) or final_payment_date

    component = {
        "sheet": sheet.title,
        "operationLabel": find_labeled_value(sheet, "Operacao"),
        "principal": round(principal, 2),
        "startDate": start_date.isoformat(),
        "accrueOnStartDate": accrue_on_start_date,
        "fundingBalance": workbook_history[-1]["balance"],
        "fundingRate": round(float(rate_terms["fundingRate"]), 4),
        "fundingRateLabel": rate_terms["fundingRateLabel"],
        "fundingRateType": rate_terms["fundingRateType"],
        "cdiRate": rate_terms["cdiRate"],
        "spreadRate": rate_terms["spreadRate"],
        "annualRate": rate_terms["annualRate"],
        "baseDays": base_days,
        "issueDate": iso_date(issue_date),
        "maturityDate": iso_date(maturity_date),
        "interestPaidToDate": round(interest_paid, 2),
        "amortizationPaidToDate": round(amortization_paid, 2),
        "installmentPaidToDate": round(installment_paid, 2),
        "events": events,
        "rateSchedule": rate_schedule,
        "workbookBalanceAtPosition": workbook_history[-1]["balance"],
    }
    component["calculatedBalanceAtPosition"] = round(calculate_component_balance(component, position_date), 2)
    component["validationDelta"] = round(component["calculatedBalanceAtPosition"] - component["workbookBalanceAtPosition"], 2)
    return component


def is_business_day(value: dt.date) -> bool:
    return value.weekday() < 5


def previous_business_day(value: dt.date) -> dt.date:
    candidate = value - dt.timedelta(days=1)
    while not is_business_day(candidate):
        candidate -= dt.timedelta(days=1)
    return candidate


def daily_rate(component: dict[str, Any]) -> float:
    base_days = int(component.get("baseDays") or 360)
    if component.get("fundingRateType") == "cdi_spread":
        cdi_rate = float(component.get("cdiRate") or 0)
        spread_rate = float(component.get("spreadRate") or 0)
        return math.pow((1 + cdi_rate) * (1 + spread_rate), 1 / base_days) - 1
    monthly_rate = float(component.get("fundingRate") or 0) / 100
    return math.pow(1 + monthly_rate, 12 / base_days) - 1


def should_accrue(component: dict[str, Any], current_date: dt.date, is_start_date: bool) -> bool:
    if int(component.get("baseDays") or 360) == 252 and not is_business_day(current_date):
        return False
    if is_start_date and not component.get("accrueOnStartDate"):
        return False
    return True


def calculate_component_balance(component: dict[str, Any], position_date: dt.date) -> float:
    start_date = dt.date.fromisoformat(str(component["startDate"]))
    if position_date < start_date:
        return 0.0

    events_by_date: dict[str, float] = {}
    for event in component.get("events", []):
        date_key = str(event.get("date") or "")
        events_by_date[date_key] = events_by_date.get(date_key, 0.0) + float(event.get("amount") or 0.0)

    rates_by_date: dict[str, float] = {}
    for row in component.get("rateSchedule", []):
        date_key = str(row.get("date") or "")
        if date_key:
            rates_by_date[date_key] = float(row.get("dailyRate") or 0.0)
    max_scheduled_rate_date = max(rates_by_date) if rates_by_date else None

    balance = float(component.get("principal") or 0.0)
    current_date = start_date
    while current_date <= position_date:
        date_key = current_date.isoformat()
        is_start_date = current_date == start_date
        scheduled_rate = rates_by_date.get(date_key)
        if scheduled_rate is not None:
            if not is_start_date or component.get("accrueOnStartDate"):
                balance += balance * scheduled_rate
        elif not max_scheduled_rate_date or date_key > max_scheduled_rate_date:
            if should_accrue(component, current_date, is_start_date):
                balance += balance * daily_rate(component)
        balance -= events_by_date.get(date_key, 0.0)
        if abs(balance) < 1:
            balance = 0.0
        current_date += dt.timedelta(days=1)
    return balance


def operation_balance_at(components: list[dict[str, Any]], position_date: dt.date) -> float:
    return sum(calculate_component_balance(component, position_date) for component in components)


def weighted_rate(components: list[dict[str, Any]]) -> float:
    active_balance = sum(float(component["calculatedBalanceAtPosition"]) for component in components)
    if active_balance:
        return sum(float(component["calculatedBalanceAtPosition"]) * float(component["fundingRate"]) for component in components) / active_balance
    principal = sum(float(component["principal"]) for component in components)
    if principal:
        return sum(float(component["principal"]) * float(component["fundingRate"]) for component in components) / principal
    return 0.0


def max_date(values: list[str | None]) -> str | None:
    clean_values = [value for value in values if value]
    return max(clean_values) if clean_values else None


def min_date(values: list[str | None]) -> str | None:
    clean_values = [value for value in values if value]
    return min(clean_values) if clean_values else None


def build_operations(workbook_path: Path, position_date: dt.date) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    operations = []

    for group in GROUPS:
        components = [sheet_position(workbook[sheet_name], position_date) for sheet_name in group["sheets"]]
        funding_balance = round(operation_balance_at(components, position_date), 2)
        funding_previous = round(operation_balance_at(components, previous_business_day(position_date)), 2)
        funding_month_start = round(operation_balance_at(components, position_date.replace(day=1)), 2)
        funding_rate = round(weighted_rate(components), 4)
        maturity_date = max_date([component["maturityDate"] for component in components])
        duration = 0
        if maturity_date:
            duration = max(0, (dt.date.fromisoformat(maturity_date) - position_date).days)

        operation = {
            "id": group["id"],
            "investor": group["investor"],
            "name": group["name"],
            "shortName": group["shortName"],
            "account": group["account"],
            "positionDate": position_date.isoformat(),
            "sourceSheets": group["sheets"],
            "issueDate": min_date([component["issueDate"] for component in components]),
            "maturityDate": maturity_date,
            "fundingPrincipal": round(sum(float(component["principal"]) for component in components), 2),
            "fundingBalance": funding_balance,
            "fundingPrevious": funding_previous,
            "fundingMonthStart": funding_month_start,
            "fundingRate": funding_rate,
            "fundingRateLabel": " / ".join(sorted({str(component["fundingRateLabel"]) for component in components})),
            "fundingRateType": "mixed" if len({component["fundingRateType"] for component in components}) > 1 else components[0]["fundingRateType"],
            "baseDays": max(component["baseDays"] for component in components),
            "cash": 0,
            "portfolioVp": 0,
            "portfolioVn": 0,
            "portfolioRate": 0,
            "duration": duration,
            "overdue": 0,
            "warning": "ok",
            "syntheticSub": round(-funding_balance, 2),
            "previousSyntheticSub": round(-funding_previous, 2),
            "monthStartSyntheticSub": round(-funding_month_start, 2),
            "portfolio": [],
            "fundingComponents": components,
            "fundingControlMode": "calculated_in_browser",
        }
        operations.append(operation)

    return operations


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_xlsx")
    parser.add_argument("--position-date")
    parser.add_argument(
        "--export-js",
        default=str(Path(__file__).resolve().parents[1] / "data" / "fundings" / "fundings-confina.js"),
    )
    args = parser.parse_args()

    source_path = Path(args.source_xlsx).resolve()
    if args.position_date:
        position_date = dt.date.fromisoformat(args.position_date)
    else:
        position_date = dt.datetime.now(dt.timezone.utc).astimezone(SAO_PAULO_TIMEZONE).date()

    operations = build_operations(source_path, position_date)
    payload = {
        "sourceWorkbook": str(source_path),
        "positionDate": position_date.isoformat(),
        "exportedAt": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "operations": operations,
    }

    export_path = Path(args.export_js)
    export_path.parent.mkdir(parents=True, exist_ok=True)
    export_path.write_text(
        f"window.{EXPORT_VARIABLE} = {json.dumps(payload, ensure_ascii=True, indent=2)};\n",
        encoding="utf-8",
    )

    total_balance = sum(operation["fundingBalance"] for operation in operations)
    print(f"{len(operations)} fundings exportados em {export_path}")
    print(f"Data da posicao: {position_date.isoformat()}")
    print(f"Saldo devedor total: {total_balance:,.2f}")
    for operation in operations:
        print(f"- {operation['shortName']}: {operation['fundingBalance']:,.2f} | {operation['fundingRate']:.4f}% a.m.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
