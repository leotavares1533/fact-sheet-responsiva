from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
CRA_ID = "cra-65"
CRA_ROOT = PROJECT_ROOT / "cras" / CRA_ID
DATA_ROOT = PROJECT_ROOT / "data" / "cras" / CRA_ID
DEFAULT_PYTHON = Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "python" / "python.exe"


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (datetime, date)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("R$", "").replace("%", "").strip()
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return 0.0
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return float(text)


def br_number(value: float) -> str:
    text = f"{float(value):.12f}".rstrip("0").rstrip(".")
    return text.replace(".", ",")


def parse_excel_date(value: object, fallback: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and value > 20000:
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date().isoformat()
    if value:
        raw = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw[:10], fmt).date().isoformat()
            except ValueError:
                pass
    return fallback


def first_value_after(row: list[object], index: int) -> object:
    for value in row[index + 1 :]:
        if value not in (None, ""):
            return value
    return 0.0


def assign_cash_value(accounts: dict[str, float], text: str, amount: float) -> tuple[bool, float]:
    provisao = 0.0
    matched = False
    if "contacorrente" in text:
        accounts["cc"] = amount
        matched = True
    elif ("aplic" in text or "apliac" in text) and "aloc" not in text:
        accounts["conta_aplicacao"] = amount
        matched = True
    elif "zeragem" in text:
        accounts["fundo_zeragem"] = amount
        matched = True
    elif text in {"saldofd", "fd"} or "fundodesp" in text:
        accounts["fundo_despesas"] = amount
        matched = True
    elif ("dev" in text and "recurso" in text) or ("devolucao" in text and "recurso" in text):
        accounts["conta_liquidacao"] = -abs(amount)
        matched = True
    elif "liquid" in text or "recurso" in text:
        accounts["conta_liquidacao"] = amount
        matched = True
    elif "provis" in text and "desp" in text:
        provisao = abs(amount)
        matched = True
    return matched, provisao


def extract_cash(cash_path: Path, date_key: str) -> tuple[dict[str, float], float]:
    workbook = openpyxl.load_workbook(cash_path, read_only=True, data_only=True)
    default_accounts = {
        "cc": 0.0,
        "conta_aplicacao": 0.0,
        "fundo_zeragem": 0.0,
        "conta_liquidacao": 0.0,
        "fundo_despesas": 0.0,
        "provisoes": 0.0,
    }

    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        date_columns = []
        for row in rows:
            for index, value in enumerate(row):
                try:
                    if parse_excel_date(value, "") == date_key:
                        date_columns.append(index)
                except Exception:
                    pass
        for date_column in dict.fromkeys(date_columns):
            accounts = dict(default_accounts)
            provisao = 0.0
            matches = 0
            for row in rows:
                label = next((value for idx, value in enumerate(row) if idx != date_column and value not in (None, "")), "")
                text = normalize(label)
                if not text:
                    continue
                amount = parse_number(row[date_column] if date_column < len(row) else 0)
                matched, row_provisao = assign_cash_value(accounts, text, amount)
                if matched:
                    matches += 1
                    provisao = row_provisao or provisao
            if matches:
                return accounts, provisao

    for worksheet in workbook.worksheets:
        accounts = dict(default_accounts)
        provisao = 0.0
        file_date = date_key
        matches = 0
        for row in worksheet.iter_rows(values_only=True):
            values = list(row)
            for index, value in enumerate(values):
                text = normalize(value)
                if not text:
                    continue
                try:
                    parsed = parse_excel_date(value, "")
                    if parsed:
                        file_date = parsed
                except Exception:
                    pass
                amount = parse_number(first_value_after(values, index))
                matched, row_provisao = assign_cash_value(accounts, text, amount)
                if matched:
                    matches += 1
                    provisao = row_provisao or provisao
        if matches and file_date == date_key:
            return accounts, provisao

    raise ValueError(f"Nao encontrei caixa para a data {date_key} em {cash_path.name}.")


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def net_accounts_for_provision(accounts: dict[str, float], provisao: float) -> dict[str, float]:
    adjusted = dict(accounts)
    remaining = abs(float(provisao or 0.0))
    if not remaining:
        return adjusted

    # CRA 65: o total do extrato ja vem liquido da provisao de despesas.
    # O caixa importado precisa seguir esse total para nao inflar a sub.
    for field in ("conta_aplicacao", "cc", "fundo_zeragem", "fundo_despesas", "conta_liquidacao"):
        if remaining <= 0:
            break
        available = max(float(adjusted.get(field) or 0.0), 0.0)
        deduction = min(available, remaining)
        adjusted[field] = float(adjusted.get(field) or 0.0) - deduction
        remaining -= deduction

    if remaining > 0:
        adjusted["conta_aplicacao"] = float(adjusted.get("conta_aplicacao") or 0.0) - remaining
    return adjusted


def write_cash_files(cash_path: Path, date_key: str) -> dict[str, float]:
    accounts, provisao = extract_cash(cash_path, date_key)
    cash_dir = CRA_ROOT / "imports" / "caixa"
    expenses_dir = CRA_ROOT / "imports" / "despesas"
    cash_dir.mkdir(parents=True, exist_ok=True)
    expenses_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(cash_path, cash_dir / "fonte-caixa.xlsx")

    raw_cash_total = sum(accounts.values())
    accounts_for_csv = net_accounts_for_provision(accounts, provisao)
    cash_total = sum(accounts_for_csv.values())
    observation = f"Total do extrato {br_number(cash_total)}"
    if provisao:
        observation = (
            f"Caixa bruto {br_number(raw_cash_total)} ajustado pela provisao de despesa de {br_number(provisao)}. "
            f"Total liquido usado no ativo {br_number(cash_total)}; provisao nao registrada como nova deducao no passivo."
        )
    write_csv(
        cash_dir / "caixa.csv",
        ["data_base", "cc", "conta_aplicacao", "fundo_zeragem", "conta_liquidacao", "fundo_despesas", "provisoes", "fonte", "arquivo_origem", "observacao"],
        [
            {
                "data_base": date_key,
                "cc": br_number(accounts_for_csv["cc"]),
                "conta_aplicacao": br_number(accounts_for_csv["conta_aplicacao"]),
                "fundo_zeragem": br_number(accounts_for_csv["fundo_zeragem"]),
                "conta_liquidacao": br_number(accounts_for_csv["conta_liquidacao"]),
                "fundo_despesas": br_number(accounts_for_csv["fundo_despesas"]),
                "provisoes": "0",
                "fonte": "Extrato Bancario",
                "arquivo_origem": cash_path.name,
                "observacao": observation,
            }
        ],
    )

    expense_rows = []
    write_csv(
        expenses_dir / "despesas.csv",
        ["data_base", "tipo", "descricao", "valor", "fonte", "arquivo_origem", "observacao"],
        expense_rows,
    )
    return {
        **accounts_for_csv,
        "provisao_despesa": provisao,
        "total_contas": cash_total,
        "total_contas_bruto": raw_cash_total,
    }


def run_command(args: list[object]) -> None:
    subprocess.run([str(item) for item in args], cwd=PROJECT_ROOT, check=True)


def write_js_snapshot(path: Path, cra_id: str, date_key: str, snapshot: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(snapshot, ensure_ascii=False, indent=2)
    path.write_text(
        "window.LAMINA_CRA_DAILY = window.LAMINA_CRA_DAILY || {};\n"
        f'window.LAMINA_CRA_DAILY["{cra_id}"] = window.LAMINA_CRA_DAILY["{cra_id}"] || {{}};\n'
        f'window.LAMINA_CRA_DAILY["{cra_id}"]["{date_key}"] = {payload};\n',
        encoding="utf-8",
    )


def load_manifest(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8-sig")
    start = text.find("[")
    end = text.rfind("]")
    if start < 0 or end < 0:
        return []
    return json.loads(text[start : end + 1])


def update_manifest_from_snapshot(snapshot: dict, date_key: str) -> None:
    manifest_path = PROJECT_ROOT / "data" / "cra-manifest.js"
    if not manifest_path.exists():
        return
    manifest = load_manifest(manifest_path)
    for cra in manifest:
        if cra.get("craId") != CRA_ID:
            continue
        cra["currentDate"] = date_key
        dates = [item for item in cra.get("dates", []) if item.get("dateKey") != date_key]
        dates.insert(
            0,
            {
                "dateKey": date_key,
                "reportDate": snapshot.get("metadata", {}).get("reportDate", ""),
                "importedAt": snapshot.get("metadata", {}).get("importedAt", ""),
                "revisionId": snapshot.get("metadata", {}).get("revisionId", ""),
                "totalAtivo": snapshot.get("ativo", {}).get("total", 0),
                "carteiraVp": snapshot.get("ativo", {}).get("carteiraVp", 0),
                "carteiraVpBruto": snapshot.get("ativo", {}).get("carteiraVpBruto", 0),
                "pddTotal": snapshot.get("ativo", {}).get("pddTotal", 0),
                "funding": snapshot.get("passivo", {}).get("fundingTotal", 0),
                "subordinada": snapshot.get("passivo", {}).get("subordinadaTotal", 0),
                "dataScript": f"data/cras/{CRA_ID}/{date_key}.js",
            },
        )
        cra["dates"] = dates
    manifest_path.write_text("window.LAMINA_CRA_MANIFEST = " + json.dumps(manifest, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def patch_cra65_pu(date_key: str) -> dict:
    canonical = CRA_ROOT / "archive" / "canonical" / f"{date_key}.json"
    if not canonical.exists():
        raise FileNotFoundError(canonical)
    snapshot = json.loads(canonical.read_text(encoding="utf-8"))
    write_js_snapshot(DATA_ROOT / f"{date_key}.js", CRA_ID, date_key, snapshot)
    write_js_snapshot(CRA_ROOT / "data" / "daily" / f"{date_key}.js", CRA_ID, date_key, snapshot)
    update_manifest_from_snapshot(snapshot, date_key)
    return snapshot


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Importa carteira e caixa do CRA 65 para uma data-base.")
    parser.add_argument("--date-key", required=True, help="Data-base yyyy-mm-dd.")
    parser.add_argument("--carteira", required=True, help="Arquivo Excel de carteira.")
    parser.add_argument("--caixa", required=True, help="Arquivo Excel de caixa.")
    parser.add_argument("--python", default=str(DEFAULT_PYTHON if DEFAULT_PYTHON.exists() else sys.executable), help="Python usado para chamar os scripts auxiliares.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    date_key = args.date_key
    carteira_path = Path(args.carteira)
    caixa_path = Path(args.caixa)
    python_path = Path(args.python)
    if not carteira_path.exists():
        raise FileNotFoundError(carteira_path)
    if not caixa_path.exists():
        raise FileNotFoundError(caixa_path)

    run_command(
        [
            python_path,
            PROJECT_ROOT / "scripts" / "update-di-rates.py",
            "--project-root",
            PROJECT_ROOT,
            "--target-date",
            date_key,
            "--soft-fail",
        ]
    )
    run_command(
        [
            python_path,
            PROJECT_ROOT / "scripts" / "normalize_carteira.py",
            "--excel",
            carteira_path,
            "--output",
            CRA_ROOT / "imports" / "carteira" / "carteira.csv",
            "--validation",
            CRA_ROOT / "imports" / "carteira" / "validacao-carteira.json",
            "--primary",
            "excel",
            "--report-date",
            date_key,
        ]
    )
    cash = write_cash_files(caixa_path, date_key)
    run_command(
        [
            python_path,
            PROJECT_ROOT / "scripts" / "refresh_cra_snapshot_from_imports.py",
            "--project-root",
            PROJECT_ROOT,
            "--cra-id",
            CRA_ID,
            "--date-key",
            date_key,
        ]
    )
    snapshot = patch_cra65_pu(date_key)
    print(
        json.dumps(
            {
                "dateKey": date_key,
                "carteira": str(carteira_path),
                "caixa": str(caixa_path),
                "linhasCarteira": len(snapshot.get("carteira", []) or []),
                "vpBruto": snapshot.get("ativo", {}).get("carteiraVpBruto", 0),
                "pdd": snapshot.get("ativo", {}).get("pddTotal", 0),
                "caixaTotalContas": cash["total_contas"],
                "provisaoDespesa": cash["provisao_despesa"],
                "ativo": snapshot.get("ativo", {}).get("total", 0),
                "funding": snapshot.get("passivo", {}).get("fundingTotal", 0),
                "subordinada": snapshot.get("passivo", {}).get("subordinadaTotal", 0),
                "puSub": snapshot.get("passivo", {}).get("subordinadaPuResidual", 0),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
