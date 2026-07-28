from __future__ import annotations

import argparse
import copy
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path.home() / "Downloads" / "carteira_caixa_ccs_2707.xlsx"
SAO_PAULO = timezone(timedelta(hours=-3))
GROUP_ID = "cras-carteira"
GROUP_NAME = "CRAs Carteira"
ACTIVE_VP_THRESHOLD = 1.0


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return 0.0
    text = text.replace("R$", "").replace("%", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    text = clean_text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def iso(value: date | None) -> str:
    return value.isoformat() if value else ""


def br_date(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def date_key_to_br(value: str) -> str:
    parsed = as_date(value)
    return br_date(parsed)


def cra_number(label: Any) -> int | None:
    match = re.search(r"\d+", clean_text(label))
    return int(match.group(0)) if match else None


def is_cra_emission_label(label: Any) -> bool:
    text = clean_text(label).lower()
    return bool(re.match(r"^\d{2}\s*(?:ª|a|º|o)?$", text))


def cra_id_from_number(number: int) -> str:
    return f"cra-carteira-{number}"


def normalize_key(text: Any) -> str:
    import unicodedata

    raw = unicodedata.normalize("NFD", clean_text(text).lower())
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", raw)


def pdd_status_from_due(base_date: date, due_date: date | None, source: str = "") -> tuple[str, int]:
    if normalize_key(source).startswith("liquid"):
        return "LIQUIDADO", 99
    if not due_date:
        return "Sem vencimento", 98
    overdue = (base_date - due_date).days
    if overdue <= 0:
        return "Em dia", 0
    if overdue <= 30:
        return "Entre 1 e 30 dias", 1
    if overdue <= 60:
        return "Entre 31 e 60 dias", 2
    if overdue <= 90:
        return "Entre 61 e 90 dias", 3
    if overdue <= 120:
        return "Entre 91 e 120 dias", 4
    if overdue <= 150:
        return "Entre 121 e 150 dias", 5
    if overdue <= 180:
        return "Entre 151 e 180 dias", 6
    return "Acima de 180 dias", 7


def pdd_rate(label: str) -> float:
    key = normalize_key(label)
    if not key or "emdia" in key or "liquid" in key:
        return 0.0
    if "1a30" in key or "31a60" in key or "entre1e30" in key or "entre31e60" in key:
        return 0.005
    if "61a90" in key or "entre61e90" in key:
        return 0.30
    if "91a120" in key or "entre91e120" in key:
        return 0.60
    if "121a150" in key or "entre121e150" in key:
        return 0.80
    if "151a180" in key or "acimade180" in key:
        return 1.0
    return 0.0


def load_json_assignment(path: Path, assignment_prefix: str | None = None) -> Any:
    text = path.read_text(encoding="utf-8")
    if assignment_prefix:
        start = text.index(assignment_prefix) + len(assignment_prefix)
        payload = text[start:].strip()
        if payload.endswith(";"):
            payload = payload[:-1].strip()
        return json.loads(payload)
    if path.name == "cra-manifest.js":
        return json.loads(text[text.index("[") : text.rfind("]") + 1])
    raise ValueError(f"Unsupported JS data file: {path}")


def snapshot_prefix(cra_id: str, date_key: str) -> str:
    return f'window.LAMINA_CRA_DAILY["{cra_id}"]["{date_key}"] ='


def load_snapshot(path: Path, cra_id: str, date_key: str) -> dict[str, Any]:
    return load_json_assignment(path, snapshot_prefix(cra_id, date_key))


def write_snapshot(path: Path, cra_id: str, date_key: str, snapshot: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(snapshot, ensure_ascii=False, indent=2)
    path.write_text(
        "window.LAMINA_CRA_DAILY = window.LAMINA_CRA_DAILY || {};\n"
        f'window.LAMINA_CRA_DAILY["{cra_id}"] = window.LAMINA_CRA_DAILY["{cra_id}"] || {{}};\n'
        f'window.LAMINA_CRA_DAILY["{cra_id}"]["{date_key}"] = {payload};\n',
        encoding="utf-8",
    )


def write_manifest(path: Path, manifest: list[dict[str, Any]]) -> None:
    payload = json.dumps(manifest, ensure_ascii=False, indent=2)
    path.write_text(f"window.LAMINA_CRA_MANIFEST = {payload};\n", encoding="utf-8")


def get_latest_snapshot_info(cra_id: str, target_date: str) -> tuple[str, dict[str, Any]] | tuple[None, None]:
    folder = REPO_ROOT / "data" / "cras" / cra_id
    if not folder.exists():
        return None, None
    candidates = sorted(
        [p.stem for p in folder.glob("*.js") if re.match(r"\d{4}-\d{2}-\d{2}$", p.stem) and p.stem <= target_date],
        reverse=True,
    )
    for date_key in candidates:
        try:
            return date_key, load_snapshot(folder / f"{date_key}.js", cra_id, date_key)
        except Exception:
            continue
    return None, None


def parse_di_rates() -> dict[str, dict[str, Any]]:
    path = REPO_ROOT / "data" / "indices" / "di.js"
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8")
    return json.loads(text[text.index("{") : text.rfind("}") + 1])


DI_RATES = parse_di_rates()


def di_daily_for(date_key: str) -> tuple[float, float, str]:
    available = [key for key in DI_RATES if key <= date_key]
    if not available:
        return 0.0, 0.0, ""
    key = sorted(available)[-1]
    daily = float(DI_RATES[key].get("taxaDia") or 0.0)
    annual = (1 + daily) ** 252 - 1 if daily > 0 else 0.0
    return daily, annual, key


def round_decimal(value: float, decimals: int) -> float:
    factor = 10**decimals
    return math.floor((float(value) + 1e-15) * factor + 0.5) / factor


def truncate_decimal(value: float, decimals: int) -> float:
    factor = 10**decimals
    return math.trunc(float(value or 0.0) * factor) / factor


def ts_daily_factor(daily_rate: float, percentual_indexador: float = 1.0) -> float:
    tdk = round_decimal(daily_rate, 8)
    return truncate_decimal(1 + (tdk * percentual_indexador), 16)


def add_ts_factor(product: float, daily_factor: float) -> float:
    return truncate_decimal(float(product or 1.0) * float(daily_factor or 1.0), 16)


def apply_ts_factor(base_pu: float, product: float) -> float:
    vne = truncate_decimal(base_pu, 8)
    factor_di = round_decimal(product, 8)
    interest = truncate_decimal(vne * (factor_di - 1), 8)
    return vne + interest


def fixed_daily_rate(cota: dict[str, Any]) -> float:
    taxa_dia = to_number(cota.get("taxaDia"))
    if taxa_dia:
        return taxa_dia
    taxa_aa = to_number(cota.get("taxaAa"))
    return (1 + taxa_aa) ** (1 / 252) - 1 if taxa_aa else 0.0


def fixed_period_factor(cota: dict[str, Any], business_days: int) -> float:
    taxa_aa = to_number(cota.get("taxaAa"))
    return round_decimal((1 + taxa_aa) ** (business_days / 252), 9) if taxa_aa else 1.0


def is_fixed_cota(cota: dict[str, Any]) -> bool:
    metodo = clean_text(cota.get("metodo")).lower()
    return "prefixado" in metodo or (not cota.get("indexador") and (to_number(cota.get("taxaAa")) or to_number(cota.get("taxaDia"))))


def extract_spread_from_text(text: str) -> float | None:
    match = re.search(r"CDI\s*\+\s*([\d.,]+)\s*%", text, flags=re.I)
    if not match:
        return None
    return to_number(match.group(1)) / 100


def cota_tax_text(snapshot: dict[str, Any], cota: dict[str, Any]) -> str:
    if_codigo = clean_text(cota.get("ifCodigo"))
    classe = clean_text(cota.get("classe")).upper()
    for row in snapshot.get("performanceCotas") or []:
        if clean_text(row.get("ifCodigo")) == if_codigo or clean_text(row.get("classe")).upper() == classe:
            if row.get("taxa"):
                return clean_text(row.get("taxa"))
    return clean_text(cota.get("taxa") or cota.get("remuneracao") or "")


def infer_spread(snapshot: dict[str, Any], cota: dict[str, Any], latest_row: dict[str, Any]) -> float:
    for key in ("spreadAa", "taxaSpreadAa", "spreadAnual"):
        value = cota.get(key)
        if value not in (None, ""):
            parsed = to_number(value)
            return parsed / 100 if parsed > 1 else parsed
    parsed = extract_spread_from_text(cota_tax_text(snapshot, cota))
    if parsed is not None:
        return parsed
    annual = to_number(latest_row.get("taxaDiAnualEquivalente"))
    _, di_annual, rate_key = di_daily_for(clean_text(latest_row.get("dataIso")) or clean_text(latest_row.get("data")) or "")
    if annual and di_annual and annual > di_annual:
        spread = (1 + annual) / (1 + di_annual) - 1
        if 0 <= spread <= 0.25:
            return spread
    return 0.0


def combined_di_spread_daily(date_key: str, snapshot: dict[str, Any], cota: dict[str, Any], latest_row: dict[str, Any]) -> tuple[float, float, str]:
    di_daily, di_annual, rate_key = di_daily_for(date_key)
    if not di_daily:
        fallback = to_number(latest_row.get("taxaDiUtilizadaDia") or latest_row.get("tdk"))
        annual = to_number(latest_row.get("taxaDiAnualEquivalente"))
        return fallback, annual, clean_text(latest_row.get("dataTaxaDiIso") or latest_row.get("dataReferenciaTaxaDiIso"))
    spread = infer_spread(snapshot, cota, latest_row)
    annual = (1 + di_annual) * (1 + spread) - 1 if spread else di_annual
    daily = (1 + annual) ** (1 / 252) - 1 if annual else di_daily
    return daily, annual, rate_key


def is_business_day(day: date, holidays: set[str]) -> bool:
    return day.weekday() < 5 and day.isoformat() not in holidays


def project_cota_to_date(snapshot: dict[str, Any], cota: dict[str, Any], target_key: str) -> dict[str, Any]:
    target = as_date(target_key)
    if not target:
        return copy.deepcopy(cota)

    next_cota = copy.deepcopy(cota)
    holidays = set(snapshot.get("metadata", {}).get("businessCalendar", {}).get("holidays") or [])
    history = copy.deepcopy(next_cota.get("historicoPu") or [])
    history = [
        row for row in history
        if clean_text(row.get("dataIso") or row.get("data")) and clean_text(row.get("dataIso") or row.get("data"))[:10] <= target_key
    ]
    history.sort(key=lambda row: clean_text(row.get("dataIso") or row.get("data"))[:10])
    if not history:
        current_key = clean_text(snapshot.get("metadata", {}).get("dateKey")) or target_key
        history = [{
            "data": date_key_to_br(current_key),
            "dataIso": current_key,
            "valorNominal": to_number(next_cota.get("principalResidual") or next_cota.get("valorNominalInicial")),
            "puAtualizado": to_number(next_cota.get("pu") or next_cota.get("principalResidual")),
            "valorReais": to_number(next_cota.get("valor")),
            "diasUteis": 0,
            "diasUteisPeriodo": 0,
            "produtorioFatorDi": 1,
            "fatorDiAcumulado": 1,
        }]

    latest = history[-1]
    latest_key = clean_text(latest.get("dataIso") or latest.get("data"))[:10]
    if "/" in latest_key:
        latest_key = iso(as_date(latest_key))
    cursor = as_date(latest_key) or target

    principal = to_number(latest.get("valorNominal") or next_cota.get("principalResidual") or next_cota.get("valorNominalInicial"))
    pu = to_number(latest.get("puAtualizado") or next_cota.get("pu") or principal)
    base_pu = pu
    period_factor = to_number(latest.get("fatorDiAcumulado")) or 1.0
    total_factor = to_number(latest.get("produtorioFatorDi")) or 1.0
    dias_uteis = int(to_number(latest.get("diasUteis")))
    dias_uteis_periodo = int(to_number(latest.get("diasUteisPeriodo") or latest.get("diasUteis")))
    pure_di_factor = 1.0
    fixed = is_fixed_cota(next_cota)
    metodo = clean_text(next_cota.get("metodo")).lower()

    while cursor < target:
        cursor += timedelta(days=1)
        day_key = cursor.isoformat()
        dia_util = is_business_day(cursor, holidays)
        fator_diario = 1.0
        taxa_dia = 0.0
        taxa_aa = 0.0
        rate_key = ""
        if dia_util and principal > 0:
            dias_uteis += 1
            dias_uteis_periodo += 1
            if fixed:
                taxa_dia = fixed_daily_rate(next_cota)
                taxa_aa = to_number(next_cota.get("taxaAa"))
                fator_diario = 1 + taxa_dia
                period_factor = fixed_period_factor(next_cota, dias_uteis_periodo)
                total_factor = fixed_period_factor(next_cota, dias_uteis)
                pure_di_factor = total_factor
                pu = apply_ts_factor(principal, period_factor)
            elif "di_spread" in metodo:
                taxa_dia, taxa_aa, rate_key = combined_di_spread_daily(day_key, snapshot, next_cota, latest)
                fator_diario = ts_daily_factor(taxa_dia, to_number(next_cota.get("percentualIndexador")) or 1.0)
                period_factor = add_ts_factor(period_factor, fator_diario)
                total_factor = add_ts_factor(total_factor, fator_diario)
                di_daily, _, _ = di_daily_for(day_key)
                pure_di_factor = add_ts_factor(pure_di_factor, ts_daily_factor(di_daily or taxa_dia, 1.0))
                pu = apply_ts_factor(base_pu, period_factor)
            else:
                taxa_dia = to_number(latest.get("taxaDiUtilizadaDia") or latest.get("tdk"))
                taxa_aa = to_number(latest.get("taxaDiAnualEquivalente"))
                fator_diario = ts_daily_factor(taxa_dia, to_number(next_cota.get("percentualIndexador")) or 1.0)
                period_factor = add_ts_factor(period_factor, fator_diario)
                total_factor = add_ts_factor(total_factor, fator_diario)
                pure_di_factor = add_ts_factor(pure_di_factor, ts_daily_factor(taxa_dia, 1.0))
                pu = apply_ts_factor(base_pu, period_factor)

        fator = pu / principal if principal else 0.0
        row = {
            "data": br_date(cursor),
            "dataIso": day_key,
            "diaUtil": dia_util,
            "taxaDiUtilizadaDia": taxa_dia or to_number(latest.get("taxaDiUtilizadaDia") or latest.get("tdk")),
            "taxaDiAnualEquivalente": taxa_aa or to_number(latest.get("taxaDiAnualEquivalente")),
            "dataTaxaDi": date_key_to_br(rate_key) if rate_key else clean_text(latest.get("dataTaxaDi")),
            "dataTaxaDiIso": rate_key or clean_text(latest.get("dataTaxaDiIso")),
            "taxaDiStatus": "projetada" if not fixed else "prefixada",
            "dataReferenciaTaxaDi": date_key_to_br(rate_key) if rate_key else clean_text(latest.get("dataReferenciaTaxaDi")),
            "dataReferenciaTaxaDiIso": rate_key or clean_text(latest.get("dataReferenciaTaxaDiIso")),
            "diasUteis": dias_uteis,
            "diasUteisPeriodo": dias_uteis_periodo,
            "fator": fator,
            "valorNominal": principal,
            "puAtualizado": pu,
            "puJuros": max(0.0, pu - principal),
            "puAntesEvento": pu,
            "puEvento": 0,
            "puAposEvento": pu,
            "principalAntesEvento": principal,
            "principalAposEvento": principal,
            "valorReais": pu * to_number(next_cota.get("quantidade")),
            "valorEventoReais": 0,
            "tdk": taxa_dia or to_number(latest.get("taxaDiUtilizadaDia") or latest.get("tdk")),
            "fatorDiario": fator_diario,
            "produtorioFatorDi": total_factor,
            "fatorDiAcumulado": period_factor,
            "spread": (to_number(next_cota.get("percentualIndexador")) or 1.0) - 1,
            "spreadAcumulado": total_factor - pure_di_factor,
            "fatorJurosAcumulado": fator,
            "evento": "",
            "eventoTs": "",
            "efeitoEvento": "",
            "ehDataPagamentoTs": False,
        }
        history.append(row)
        latest = row

    final_row = history[-1]
    next_cota["historicoPu"] = history
    next_cota["pu"] = to_number(final_row.get("puAtualizado"))
    next_cota["principalResidual"] = to_number(final_row.get("valorNominal"))
    next_cota["valor"] = to_number(final_row.get("valorReais")) or next_cota["pu"] * to_number(next_cota.get("quantidade"))
    next_cota["dataHistoricaDisponivel"] = True
    next_cota["dataHistoricaSelecionada"] = final_row.get("data")
    next_cota["dataHistoricaIso"] = final_row.get("dataIso")
    next_cota["visaoCalculadaPu"] = final_row.get("dataIso") == target_key
    next_cota["acumulacaoFinal"] = {
        **(next_cota.get("acumulacaoFinal") or {}),
        "periodoFim": final_row.get("data"),
        "diasAcumulacao": to_number(final_row.get("diasUteis")),
        "diasUteisPeriodo": to_number(final_row.get("diasUteisPeriodo")),
        "puFinal": next_cota["pu"],
    }
    return next_cota


def parse_workbook(path: Path) -> tuple[str, dict[str, list[dict[str, Any]]], dict[str, dict[str, Any]], dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = next(name for name in wb.sheetnames if normalize_key(name).startswith("carteira"))
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[2]]
    header_index = {header: index for index, header in enumerate(headers) if header is not None}
    date_header = headers[-1]
    base_date = as_date(date_header)
    if not base_date:
        raise ValueError("Nao consegui identificar a data-base na ultima coluna da carteira.")
    vp_index = len(headers) - 1
    date_key = iso(base_date)
    by_cra: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for raw in ws.iter_rows(min_row=3, values_only=True):
        number = cra_number(raw[header_index["CRA Carteira"]])
        if not number:
            continue
        cra_id = cra_id_from_number(number)
        due_date = as_date(raw[header_index["DT Vencimento"]])
        acquisition_date = as_date(raw[header_index["DT Aquisição"]])
        liquidation_date = as_date(raw[header_index["Dt Liquidação"]])
        vp_bruto = to_number(raw[vp_index])
        source_status = clean_text(raw[header_index["Status"]])
        source_payment_status = clean_text(raw[header_index["Status de Pagamento"]])
        source_faixa = clean_text(raw[header_index["Faixa Venc"]])
        faixa, faixa_order = pdd_status_from_due(base_date, due_date, source_faixa)
        if source_faixa and normalize_key(source_faixa) not in ("liquidado",):
            faixa = source_faixa
        pdd = 0.0
        is_active = vp_bruto > ACTIVE_VP_THRESHOLD
        tx_op = to_number(raw[header_index["Tx Op"]])
        item = {
            "craCarteira": clean_text(raw[header_index["CRA Carteira"]]),
            "lastro": clean_text(raw[header_index["Lastro"]]),
            "numeroUnico": clean_text(raw[header_index["Lastro"]]),
            "cedente": clean_text(raw[header_index["Cedente"]]) or "Nao informado",
            "sacado": clean_text(raw[header_index["Sacado"]]) or "Nao informado",
            "devedor": clean_text(raw[header_index["Sacado"]]) or "Nao informado",
            "valorAquisicao": to_number(raw[header_index["Vlr Aquisição"]]),
            "valorFace": to_number(raw[header_index["Vlr Face"]]),
            "valorNominal": to_number(raw[header_index["Vlr Face"]]),
            "dataAquisicao": br_date(acquisition_date),
            "dataAquisicaoIso": iso(acquisition_date),
            "dataVencimento": br_date(due_date),
            "dataVencimentoIso": iso(due_date),
            "valorLiquidacao": to_number(raw[header_index["Vlr Liquidação"]]),
            "dataLiquidacao": "" if is_active else br_date(liquidation_date),
            "dataLiquidacaoIso": "" if is_active else iso(liquidation_date),
            "dataLiquidacaoOriginal": br_date(liquidation_date),
            "dataLiquidacaoOriginalIso": iso(liquidation_date),
            "tipoTitulo": clean_text(raw[header_index["tipo_titulo"]]) or "-",
            "tipoAtivo": clean_text(raw[header_index["tipo_titulo"]]) or "-",
            "mesaResp": clean_text(raw[header_index["Mesa Resp"]]),
            "status": "EM CARTEIRA" if is_active else source_status,
            "statusOriginal": source_status,
            "statusPagamento": source_payment_status,
            "taxa": tx_op,
            "taxaCessao": to_number(raw[header_index["Tx Cessão"]]),
            "taxaOp": tx_op,
            "taxaMedia": tx_op,
            "base": to_number(raw[header_index["Base"]]),
            "tipo": clean_text(raw[header_index["Tipo"]]),
            "indexadorAtivo": clean_text(raw[header_index["Tipo"]]),
            "pagamento": clean_text(raw[header_index["Pagamento"]]),
            "dias": to_number(raw[header_index["Dias"]]),
            "diasDoVencimento": to_number(raw[header_index["Dias do Venc."]]),
            "diasUteisOp": to_number(raw[header_index["Dias Úteis Op."]]),
            "diasCorridosOp": to_number(raw[header_index["Dias Corridos Op."]]),
            "fator": to_number(raw[header_index["Fator"]]),
            "faixaVenc": faixa,
            "faixaVencOrder": faixa_order,
            "valorPresenteDia": vp_bruto,
            "valorPresente": vp_bruto,
            "pdd": pdd,
            "valorPresenteLiquido": vp_bruto,
        }
        by_cra[cra_id].append(item)

    cash_sheet = wb["Planilha2"] if "Planilha2" in wb.sheetnames else wb[wb.sheetnames[1]]
    cash_by_cra: dict[str, dict[str, Any]] = {}
    current_id = ""
    current_accounts: dict[str, float] = {}
    current_raw: list[dict[str, Any]] = []
    cash_date = base_date
    for row in cash_sheet.iter_rows(values_only=True):
        label = clean_text(row[1] if len(row) > 1 else "")
        desc = clean_text(row[5] if len(row) > 5 else "")
        value = to_number(row[6] if len(row) > 6 else 0)
        maybe_date = as_date(row[6] if len(row) > 6 else None)
        if maybe_date:
            cash_date = maybe_date
        if is_cra_emission_label(label):
            if current_id:
                cash_by_cra[current_id] = build_cash(current_accounts, current_raw)
            current_id = cra_id_from_number(cra_number(label))
            current_accounts = {}
            current_raw = []
        if not current_id:
            continue
        key = normalize_key(desc)
        if "contacorrente" in key:
            current_accounts["cc"] = value
        elif "aplicacao" in key:
            current_accounts["contaAplicacao"] = value
        elif "fundodedespesa" in key:
            current_accounts["fundoDespesas"] = value
        elif key == "total":
            current_accounts["totalInformado"] = value
        if desc:
            current_raw.append({"label": desc, "valor": value})
    if current_id:
        cash_by_cra[current_id] = build_cash(current_accounts, current_raw)

    meta = {
        "sourceFile": str(path),
        "sheetCarteira": sheet_name,
        "sheetCaixa": cash_sheet.title,
        "cashDate": iso(cash_date),
    }
    return date_key, by_cra, cash_by_cra, meta


def build_cash(accounts: dict[str, float], rows: list[dict[str, Any]]) -> dict[str, Any]:
    cc = to_number(accounts.get("cc"))
    aplicacao = to_number(accounts.get("contaAplicacao"))
    fundo_despesas = to_number(accounts.get("fundoDespesas"))
    total_calc = cc + aplicacao - fundo_despesas
    total = accounts.get("totalInformado")
    if total is None:
        total = max(total_calc, 0.0)
    return {
        "accounts": {
            "cc": cc,
            "contaAplicacao": aplicacao,
            "fundoDespesas": fundo_despesas,
        },
        "total": max(0.0, to_number(total)),
        "totalCalculado": total_calc,
        "rawRows": rows,
    }


def weighted_average(rows: list[dict[str, Any]], value_key: str, weight_key: str) -> float:
    total_weight = sum(max(0.0, to_number(row.get(weight_key))) for row in rows)
    if not total_weight:
        return 0.0
    return sum(to_number(row.get(value_key)) * max(0.0, to_number(row.get(weight_key))) for row in rows) / total_weight


def group_sum(rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        name = clean_text(row.get(key)) or "Nao informado"
        current = groups.setdefault(name, {
            "name": name,
            "quantidade": 0,
            "valorPresenteDia": 0.0,
            "pdd": 0.0,
            "valorPresenteLiquido": 0.0,
            "valorNominal": 0.0,
        })
        current["quantidade"] += 1
        current["valorPresenteDia"] += to_number(row.get("valorPresenteDia"))
        current["pdd"] += to_number(row.get("pdd"))
        current["valorPresenteLiquido"] += to_number(row.get("valorPresenteLiquido"))
        current["valorNominal"] += to_number(row.get("valorNominal"))
    total = sum(row["valorPresenteLiquido"] for row in groups.values())
    return [
        {**row, "posicao": index + 1, "participacao": row["valorPresenteLiquido"] / total if total else 0.0, "participacaoPl": row["valorPresenteLiquido"] / total if total else 0.0}
        for index, row in enumerate(sorted(groups.values(), key=lambda item: item["valorPresenteLiquido"], reverse=True))
    ]


def build_portfolio_blocks(rows: list[dict[str, Any]], base_date: date, ativo_total: float) -> dict[str, Any]:
    active = [row for row in rows if to_number(row.get("valorPresenteDia")) > ACTIVE_VP_THRESHOLD]
    vp_bruto = sum(to_number(row.get("valorPresenteDia")) for row in active)
    pdd_total = sum(to_number(row.get("pdd")) for row in active)
    vp_liquido = sum(to_number(row.get("valorPresenteLiquido")) for row in active)
    valor_nominal = sum(to_number(row.get("valorNominal")) for row in active)
    prazo = weighted_average(active, "diasDoVencimento", "valorPresenteLiquido")
    taxa_media = weighted_average(active, "taxaOp", "valorPresenteLiquido")
    pre = [row for row in active if "pre" in normalize_key(row.get("tipo"))]
    pos = [row for row in active if "pos" in normalize_key(row.get("tipo"))]
    overdue = [
        row for row in active
        if as_date(row.get("dataVencimentoIso")) and as_date(row.get("dataVencimentoIso")) < base_date
    ]

    by_cedente = group_sum(active, "cedente")
    by_sacado = group_sum(active, "sacado")
    base_pl = ativo_total if ativo_total else vp_liquido
    for ranking in (by_cedente, by_sacado):
        for row in ranking:
            row["participacaoPl"] = row["valorPresenteLiquido"] / base_pl if base_pl else 0.0

    aging_groups: dict[str, dict[str, Any]] = {}
    for row in active:
        label = row.get("faixaVenc") or "Sem faixa"
        current = aging_groups.setdefault(label, {
            "status": label,
            "valorNominal": 0.0,
            "valorPresente": 0.0,
            "valorPdd": 0.0,
            "order": to_number(row.get("faixaVencOrder")),
        })
        current["valorNominal"] += to_number(row.get("valorNominal"))
        current["valorPresente"] += to_number(row.get("valorPresenteDia"))
        current["valorPdd"] += to_number(row.get("pdd"))
    aging = []
    for item in aging_groups.values():
        item["percentualCarteira"] = item["valorPresente"] / vp_bruto if vp_bruto else 0.0
        aging.append(item)
    aging.sort(key=lambda item: item.get("order", 99))

    pdd_by_cedente_status: dict[tuple[str, str], dict[str, Any]] = {}
    for row in active:
        if to_number(row.get("pdd")) <= 0:
            continue
        key = (row["cedente"], row["faixaVenc"])
        current = pdd_by_cedente_status.setdefault(key, {
            "cedente": row["cedente"],
            "status": row["faixaVenc"],
            "valorAberto": 0.0,
            "valorPdd": 0.0,
            "valorNominal": 0.0,
            "lastros": 0,
        })
        current["valorAberto"] += to_number(row.get("valorPresenteDia"))
        current["valorPdd"] += to_number(row.get("pdd"))
        current["valorNominal"] += to_number(row.get("valorNominal"))
        current["lastros"] += 1

    composition_by_type: dict[str, dict[str, Any]] = {}
    for row in active:
        label = row.get("tipoTitulo") or "-"
        current = composition_by_type.setdefault(label, {
            "label": label,
            "tipoTitulo": label,
            "valorNominal": 0.0,
            "valorPresente": 0.0,
            "lastros": 0,
        })
        current["valorNominal"] += to_number(row.get("valorNominal"))
        current["valorPresente"] += to_number(row.get("valorPresenteDia"))
        current["lastros"] += 1
    composition = []
    for row in composition_by_type.values():
        row["percentualCarteira"] = row["valorPresente"] / vp_bruto if vp_bruto else 0.0
        row["taxaMedia"] = weighted_average([item for item in active if item.get("tipoTitulo") == row["tipoTitulo"]], "taxaOp", "valorPresenteLiquido")
        composition.append(row)
    composition.sort(key=lambda item: item["valorPresente"], reverse=True)

    acquisitions = [row for row in active if row.get("dataAquisicaoIso") == iso(base_date)]
    liquidations = [row for row in rows if row.get("dataLiquidacaoOriginalIso") == iso(base_date) and to_number(row.get("valorLiquidacao")) > 0]

    return {
        "active": active,
        "resumo": {
            "valorNominal": valor_nominal,
            "valorPresente": vp_bruto,
            "valorPresenteLiquido": vp_liquido,
            "pddTotal": pdd_total,
            "cedentesUnicos": len({row["cedente"] for row in active}),
            "sacadosUnicos": len({row["sacado"] for row in active}),
            "prazoMedioDias": prazo,
            "taxaMediaPonderada": taxa_media,
            "preFixado": {"valorPresente": sum(to_number(row.get("valorPresenteDia")) for row in pre)},
            "posFixado": {"valorPresente": sum(to_number(row.get("valorPresenteDia")) for row in pos)},
            "montanteAtraso": sum(to_number(row.get("valorPresenteDia")) for row in overdue),
        },
        "ranking": {
            "cedentes": by_cedente[:20],
            "sacados": by_sacado[:20],
            "lastrosAtivos": len(active),
        },
        "concentracaoDetalhada": {
            "top10Cedentes": by_cedente[:10],
            "top10Sacados": by_sacado[:10],
        },
        "agingList": aging,
        "pddComposition": sorted(pdd_by_cedente_status.values(), key=lambda item: item["valorPdd"], reverse=True)[:20],
        "composicaoCarteira": composition,
        "movimentacoesDia": {
            "aquisicoes": summarize_movements(acquisitions, "cedente", "valorAquisicao", base_pl),
            "liquidacoes": summarize_movements(liquidations, "cedente", "valorLiquidacao", base_pl),
        },
        "counts": {
            "active": len(active),
            "overdue": len(overdue),
            "total": len(rows),
        },
    }


def summarize_movements(rows: list[dict[str, Any]], key: str, value_key: str, base_pl: float) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        name = clean_text(row.get(key)) or "Nao informado"
        current = grouped.setdefault(name, {
            "name": name,
            "cedente": name,
            "valorAquisicao": 0.0,
            "valorFace": 0.0,
            "valorLiquidacao": 0.0,
            "quantidade": 0,
        })
        current["valorAquisicao"] += to_number(row.get("valorAquisicao"))
        current["valorFace"] += to_number(row.get("valorFace"))
        current["valorLiquidacao"] += to_number(row.get("valorLiquidacao"))
        current["quantidade"] += 1
    output = []
    for item in grouped.values():
        value = to_number(item.get(value_key))
        item["percentualPl"] = value / base_pl if base_pl else 0.0
        output.append(item)
    return sorted(output, key=lambda item: to_number(item.get(value_key)), reverse=True)[:20]


def find_prev_snapshot(cra_id: str, date_key: str) -> dict[str, Any] | None:
    folder = REPO_ROOT / "data" / "cras" / cra_id
    candidates = sorted(
        [p.stem for p in folder.glob("*.js") if re.match(r"\d{4}-\d{2}-\d{2}$", p.stem) and p.stem < date_key],
        reverse=True,
    )
    for candidate in candidates:
        try:
            return load_snapshot(folder / f"{candidate}.js", cra_id, candidate)
        except Exception:
            pass
    return None


def cota_key(cota: dict[str, Any]) -> str:
    return clean_text(cota.get("ifCodigo")) or clean_text(cota.get("classe")).upper()


def performance_rows(snapshot: dict[str, Any], prev_snapshot: dict[str, Any] | None, date_key: str) -> list[dict[str, Any]]:
    prev_by_key = {}
    for cota in (prev_snapshot or {}).get("passivo", {}).get("cotas", []) or []:
        prev_by_key[cota_key(cota)] = cota
    prev_has_asset = to_number((prev_snapshot or {}).get("ativo", {}).get("carteiraVpBruto")) > 0
    rows = []
    for cota in snapshot.get("passivo", {}).get("cotas", []) or []:
        key = cota_key(cota)
        prev = prev_by_key.get(key)
        pu = to_number(cota.get("pu"))
        prev_pu = to_number(prev.get("pu")) if prev else 0.0
        is_sub = clean_text(cota.get("classe")).upper() == "SUB" or clean_text(cota.get("tipo")).lower() == "sub"
        resultado_dia = (pu / prev_pu - 1) if prev_pu and (prev_has_asset or not is_sub) else None
        row = {
            "classe": cota.get("classe"),
            "label": cota.get("label") or cota.get("classe"),
            "tipo": cota.get("tipo"),
            "ifCodigo": cota.get("ifCodigo"),
            "quantidade": cota.get("quantidade"),
            "taxa": cota_tax_text(snapshot, cota) or cota.get("taxa") or "-",
            "pu": pu,
            "valor": to_number(cota.get("valor")),
            "resultadoDia": resultado_dia,
            "resultadoMes": None,
            "resultado30Dias": None,
            "resultadoInicio": None,
        }
        rows.append(row)
    return rows


def daily_return_history(snapshot: dict[str, Any], date_key: str) -> list[dict[str, Any]]:
    rows = []
    folder = REPO_ROOT / "data" / "cras" / snapshot["cra"]["id"]
    dates = sorted(
        [p.stem for p in folder.glob("*.js") if re.match(r"\d{4}-\d{2}-\d{2}$", p.stem) and p.stem < date_key],
        reverse=True,
    )[:29]
    dates = sorted(dates)
    historical_snapshots = []
    for key in dates:
        try:
            historical_snapshots.append(load_snapshot(folder / f"{key}.js", snapshot["cra"]["id"], key))
        except Exception:
            pass
    historical_snapshots.append(snapshot)
    previous = None
    for snap in historical_snapshots:
        cotas = {}
        prev_by_key = {cota_key(c): c for c in (previous or {}).get("passivo", {}).get("cotas", []) or []}
        prev_has_asset = to_number((previous or {}).get("ativo", {}).get("carteiraVpBruto")) > 0
        for cota in snap.get("passivo", {}).get("cotas", []) or []:
            key = cota_key(cota)
            pu = to_number(cota.get("pu"))
            prev_pu = to_number((prev_by_key.get(key) or {}).get("pu"))
            is_sub = clean_text(cota.get("classe")).upper() == "SUB" or clean_text(cota.get("tipo")).lower() == "sub"
            cotas[clean_text(cota.get("classe")).upper()] = {
                "pu": pu,
                "valor": to_number(cota.get("valor")),
                "resultadoDia": (pu / prev_pu - 1) if prev_pu and (prev_has_asset or not is_sub) else None,
                "resultadoMes": None,
            }
        rows.append({
            "dateKey": snap.get("metadata", {}).get("dateKey"),
            "reportDate": snap.get("metadata", {}).get("reportDate"),
            "cotas": cotas,
        })
        previous = snap
    return rows[-30:]


def build_sections(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    ativo = snapshot["ativo"]
    passivo = snapshot["passivo"]
    resumo = snapshot["carteiraResumo"]
    return [{
        "id": "resumo",
        "label": "Resumo",
        "metrics": [
            {"label": "Ativo total", "value": currency(ativo["total"]), "isHighlight": True, "source": {"name": "Carteira + caixa"}},
            {"label": "Carteira VP liquida", "value": currency(ativo["carteiraVpLiquido"]), "isHighlight": True, "source": {"name": "Import carteira"}},
            {"label": "Caixa total", "value": currency(ativo["caixa"]), "isHighlight": True, "source": {"name": "Import caixa"}},
            {"label": "Funding SR/MEZ", "value": currency(passivo["fundingTotal"]), "isHighlight": True, "source": {"name": "Memoria PU"}},
            {"label": "Subordinada residual", "value": currency(passivo["subordinadaTotal"]), "isHighlight": True, "source": {"name": "Residual"}},
            {"label": "PU SUB residual", "value": number(passivo["subordinadaPuResidual"], 6), "isHighlight": True, "source": {"name": "Residual"}},
            {"label": "Valor nominal", "value": currency(resumo["valorNominal"]), "isHighlight": False, "source": {"name": "Carteira"}},
            {"label": "PDD", "value": currency(ativo["pddTotal"]), "isHighlight": False, "source": {"name": "Faixa vencimento"}},
        ],
    }]


def currency(value: float) -> str:
    return "R$ " + f"{to_number(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def number(value: float, decimals: int = 2) -> str:
    return f"{to_number(value):,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def build_snapshot(cra_id: str, date_key: str, rows: list[dict[str, Any]], cash: dict[str, Any], source_meta: dict[str, Any]) -> dict[str, Any]:
    source_date_key, source_snapshot = get_latest_snapshot_info(cra_id, date_key)
    if not source_snapshot:
        number = cra_number(cra_id)
        source_snapshot = {
            "metadata": {},
            "cra": {"id": cra_id, "name": f"CRA Carteira {number}"},
            "passivo": {"cotas": [], "fundingTotal": 0, "subordinadaQuantidade": 100},
        }
    snapshot = copy.deepcopy(source_snapshot)
    base_date = as_date(date_key)
    report_date = br_date(base_date)
    cotas = [project_cota_to_date(source_snapshot, cota, date_key) for cota in source_snapshot.get("passivo", {}).get("cotas", [])]
    funding_total = sum(
        to_number(cota.get("valor"))
        for cota in cotas
        if clean_text(cota.get("classe")).upper() != "SUB" and clean_text(cota.get("tipo")).lower() != "sub"
    )
    temp_ativo_total = sum(to_number(row.get("valorPresenteLiquido")) for row in rows) + to_number(cash.get("total"))
    blocks = build_portfolio_blocks(rows, base_date, temp_ativo_total)
    resumo = blocks["resumo"]
    carteira_vp_bruto = resumo["valorPresente"]
    pdd_total = resumo["pddTotal"]
    carteira_vp_liquido = resumo["valorPresenteLiquido"]
    caixa_total = to_number(cash.get("total"))
    ativo_total = carteira_vp_liquido + caixa_total
    sub_quantity = to_number(source_snapshot.get("passivo", {}).get("subordinadaQuantidade")) or 100.0
    sub_total = ativo_total - funding_total
    sub_pu = sub_total / sub_quantity if sub_quantity else 0.0

    sub_found = False
    for cota in cotas:
        if clean_text(cota.get("classe")).upper() == "SUB" or clean_text(cota.get("tipo")).lower() == "sub":
            cota["quantidade"] = sub_quantity
            cota["pu"] = sub_pu
            cota["valor"] = sub_total
            cota["principalResidual"] = sub_pu
            sub_found = True
    if not sub_found:
        cotas.append({
            "classe": "SUB",
            "label": "Subordinada",
            "tipo": "sub",
            "ifCodigo": "",
            "quantidade": sub_quantity,
            "pu": sub_pu,
            "valor": sub_total,
            "principalResidual": sub_pu,
            "ehFunding": False,
            "ordem": 90,
        })

    snapshot["metadata"] = {
        **(snapshot.get("metadata") or {}),
        "schemaVersion": 2,
        "model": "cra-carteira",
        "craId": cra_id,
        "reportDate": report_date,
        "dateKey": date_key,
        "importedAt": datetime.now(SAO_PAULO).isoformat(),
        "revisionId": f"{date_key.replace('-', '')}-cras-carteira-consolidado",
        "sourceSnapshotDateKey": source_date_key,
        "assetImport": {
            "sourceFile": source_meta["sourceFile"],
            "sheetCarteira": source_meta["sheetCarteira"],
            "sheetCaixa": source_meta["sheetCaixa"],
            "cashDate": source_meta["cashDate"],
            "importMode": "carteira-caixa-consolidado",
        },
    }
    snapshot["cra"] = {
        **(snapshot.get("cra") or {}),
        "id": cra_id,
        "name": snapshot.get("cra", {}).get("name") or f"CRA Carteira {cra_number(cra_id)}",
        "dataBase": report_date,
        "dateKey": date_key,
    }
    snapshot["ativo"] = {
        "carteiraVpBruto": carteira_vp_bruto,
        "pddTotal": pdd_total,
        "carteiraVp": carteira_vp_liquido,
        "carteiraVpLiquido": carteira_vp_liquido,
        "caixa": caixa_total,
        "total": ativo_total,
        "liquidacoesDia": sum(to_number(row.get("valorLiquidacao")) for row in blocks["movimentacoesDia"]["liquidacoes"]),
        "aquisicoesDia": sum(to_number(row.get("valorAquisicao")) for row in blocks["movimentacoesDia"]["aquisicoes"]),
        "quantidadeLastros": len(rows),
        "quantidadeLastrosBase": len(rows),
        "lastrosAtivos": blocks["counts"]["active"],
        "quantidadeLastrosAtivos": blocks["counts"]["active"],
        "cedentesUnicos": resumo["cedentesUnicos"],
        "sacadosUnicos": resumo["sacadosUnicos"],
        "montanteAtraso": resumo["montanteAtraso"],
    }
    snapshot["caixa"] = {
        "accounts": cash.get("accounts") or {},
        "total": caixa_total,
        "totalCalculado": cash.get("totalCalculado"),
        "fonte": "Importacao consolidada CRAs Carteira",
        "arquivoOrigem": Path(source_meta["sourceFile"]).name,
        "observacao": "Fundo de despesa reduz o caixa conforme total informado na aba Planilha2.",
    }
    snapshot["passivo"] = {
        **(snapshot.get("passivo") or {}),
        "fundingTotal": funding_total,
        "despesasTotal": 0,
        "despesasOperacionaisTotal": 0,
        "provisoesTotal": pdd_total,
        "provisoesCaixa": 0,
        "deducoesTotal": 0,
        "subordinadaTotal": sub_total,
        "subordinadaQuantidade": sub_quantity,
        "subordinadaPuResidual": sub_pu,
        "cotas": cotas,
    }
    snapshot["carteira"] = rows
    snapshot["carteiraResumo"] = resumo
    snapshot["rankingCarteira"] = blocks["ranking"]
    snapshot["concentracaoDetalhada"] = blocks["concentracaoDetalhada"]
    snapshot["agingList"] = blocks["agingList"]
    snapshot["pddComposition"] = blocks["pddComposition"]
    snapshot["composicaoCarteira"] = blocks["composicaoCarteira"]
    snapshot["movimentacoesDia"] = blocks["movimentacoesDia"]
    snapshot["aquisicoesDia"] = {
        "total": snapshot["ativo"]["aquisicoesDia"],
        "quantidade": len(blocks["movimentacoesDia"]["aquisicoes"]),
        "itens": blocks["movimentacoesDia"]["aquisicoes"],
    }
    snapshot["liquidacoesDia"] = {
        "total": snapshot["ativo"]["liquidacoesDia"],
        "quantidade": len(blocks["movimentacoesDia"]["liquidacoes"]),
        "itens": blocks["movimentacoesDia"]["liquidacoes"],
    }
    prev_snapshot = find_prev_snapshot(cra_id, date_key)
    snapshot["performanceCotas"] = performance_rows(snapshot, prev_snapshot, date_key)
    snapshot["rendimento30Dias"] = daily_return_history(snapshot, date_key)
    snapshot["precificacaoMensal"] = snapshot.get("precificacaoMensal") or []
    snapshot["sections"] = build_sections(snapshot)
    snapshot["sources"] = [
        {"name": "Importacao carteira+caixa CRAs Carteira", "date": report_date, "file": Path(source_meta["sourceFile"]).name},
        {"name": "Memoria PU existente", "date": date_key_to_br(source_date_key) if source_date_key else "", "file": f"data/cras/{cra_id}/{source_date_key}.js" if source_date_key else ""},
    ]
    return snapshot


def build_overview(date_key: str, snapshots: dict[str, dict[str, Any]], source_meta: dict[str, Any]) -> dict[str, Any]:
    report_date = date_key_to_br(date_key)
    asset_rows = []
    series_rows = []
    for cra_id, snapshot in sorted(snapshots.items(), key=lambda item: cra_number(item[0]) or 0):
        number = cra_number(cra_id)
        ativo = snapshot.get("ativo", {})
        passivo = snapshot.get("passivo", {})
        resumo = snapshot.get("carteiraResumo", {})
        sub_perf = next((row for row in snapshot.get("performanceCotas", []) if clean_text(row.get("classe")).upper() == "SUB"), {})
        asset_rows.append({
            "operacao": f"CRA {number}",
            "craId": cra_id,
            "carteiraVp": ativo.get("carteiraVpLiquido"),
            "valorNominal": resumo.get("valorNominal"),
            "caixa": ativo.get("caixa"),
            "ativoTotal": ativo.get("total"),
            "funding": passivo.get("fundingTotal"),
            "subordinada": passivo.get("subordinadaTotal"),
            "puSub": passivo.get("subordinadaPuResidual"),
            "rendimentoSubDia": sub_perf.get("resultadoDia"),
            "rendimentoSubMes": sub_perf.get("resultadoMes"),
            "pdd": ativo.get("pddTotal"),
            "lastrosAtivos": ativo.get("lastrosAtivos"),
            "cedentes": resumo.get("cedentesUnicos"),
            "sacados": resumo.get("sacadosUnicos"),
            "taxaMedia": resumo.get("taxaMediaPonderada"),
            "prazoMedio": resumo.get("prazoMedioDias"),
            "montanteAtraso": resumo.get("montanteAtraso"),
        })
        for cota in passivo.get("cotas", []) or []:
            if clean_text(cota.get("classe")).upper() == "SUB" or clean_text(cota.get("tipo")).lower() == "sub":
                continue
            series_rows.append({
                "operacao": f"CRA {number}",
                "craId": cra_id,
                "serie": cota.get("label") or cota.get("classe"),
                "classe": cota.get("classe"),
                "ifCodigo": cota.get("ifCodigo"),
                "puAtual": cota.get("pu"),
                "valorAtual": cota.get("valor"),
                "quantidadeIntegralizada": cota.get("quantidade"),
                "taxa": cota_tax_text(snapshot, cota) or "-",
                "dataVencimento": snapshot.get("cra", {}).get("dataVencimento"),
                "dataVencimentoIso": snapshot.get("cra", {}).get("dataVencimentoIso"),
                "status": "Integralizada" if to_number(cota.get("quantidade")) > 0 else "Nao integralizada",
            })

    totals = {
        "carteiraVp": sum(to_number(row.get("carteiraVp")) for row in asset_rows),
        "valorNominal": sum(to_number(row.get("valorNominal")) for row in asset_rows),
        "caixa": sum(to_number(row.get("caixa")) for row in asset_rows),
        "ativoTotal": sum(to_number(row.get("ativoTotal")) for row in asset_rows),
        "funding": sum(to_number(row.get("funding")) for row in asset_rows),
        "subordinada": sum(to_number(row.get("subordinada")) for row in asset_rows),
        "pdd": sum(to_number(row.get("pdd")) for row in asset_rows),
        "lastrosAtivos": sum(to_number(row.get("lastrosAtivos")) for row in asset_rows),
    }
    history_by_date: dict[str, dict[str, Any]] = {}
    overview_folder = REPO_ROOT / "data" / "cras" / "cras-carteira-overview"
    for path in sorted(overview_folder.glob("*.js")):
        key = path.stem
        if not re.match(r"\d{4}-\d{2}-\d{2}$", key) or key >= date_key:
            continue
        try:
            previous_overview = load_snapshot(path, "cras-carteira-overview", key)
        except Exception:
            continue
        previous_history = previous_overview.get("portfolioOverview", {}).get("subHistory") or []
        for row in previous_history:
            row_key = clean_text(row.get("dateKey"))
            if row_key:
                history_by_date[row_key] = row
        previous_asset_rows = previous_overview.get("portfolioOverview", {}).get("assetRows") or []
        if previous_asset_rows:
            history_by_date[key] = {
                "dateKey": key,
                "reportDate": previous_overview.get("metadata", {}).get("reportDate") or date_key_to_br(key),
                "caixaTotal": sum(to_number(row.get("caixa")) for row in previous_asset_rows),
                "subordinadaTotal": sum(to_number(row.get("subordinada")) for row in previous_asset_rows),
                "rendimentoSubDia": None,
            }
    ordered_history = [history_by_date[key] for key in sorted(history_by_date)]
    previous_sub = to_number(ordered_history[-1].get("subordinadaTotal")) if ordered_history else 0.0
    current_sub_return = totals["subordinada"] / previous_sub - 1 if previous_sub else None
    sub_history = ordered_history[-29:] + [{
        "dateKey": date_key,
        "reportDate": report_date,
        "caixaTotal": totals["caixa"],
        "subordinadaTotal": totals["subordinada"],
        "rendimentoSubDia": current_sub_return,
    }]
    sub_quantity = 1.0
    overview = {
        "metadata": {
            "schemaVersion": 2,
            "model": "cras-carteira-consolidado",
            "craId": "cras-carteira-overview",
            "reportDate": report_date,
            "dateKey": date_key,
            "importedAt": datetime.now(SAO_PAULO).isoformat(),
            "revisionId": f"{date_key.replace('-', '')}-cras-carteira-overview-assets",
            "portfolioOverview": True,
            "portfolioConsolidado": True,
            "assetImport": {
                "sourceFile": source_meta["sourceFile"],
                "sheetCarteira": source_meta["sheetCarteira"],
                "sheetCaixa": source_meta["sheetCaixa"],
                "importMode": "carteira-caixa-consolidado",
            },
        },
        "cra": {
            "id": "cras-carteira-overview",
            "name": "CRAs Carteira - Visao geral",
            "description": "Resumo consolidado de carteira, caixa, funding e subordinada dos CRAs Carteira.",
        },
        "ativo": {
            "carteiraVpBruto": totals["carteiraVp"] + totals["pdd"],
            "pddTotal": totals["pdd"],
            "carteiraVp": totals["carteiraVp"],
            "carteiraVpLiquido": totals["carteiraVp"],
            "caixa": totals["caixa"],
            "total": totals["ativoTotal"],
            "lastrosAtivos": totals["lastrosAtivos"],
            "quantidadeLastrosAtivos": totals["lastrosAtivos"],
        },
        "caixa": {
            "accounts": {"totalCrasCarteira": totals["caixa"]},
            "total": totals["caixa"],
            "fonte": "Soma dos caixas importados por CRA",
            "arquivoOrigem": Path(source_meta["sourceFile"]).name,
        },
        "passivo": {
            "fundingTotal": totals["funding"],
            "subordinadaTotal": totals["subordinada"],
            "subordinadaQuantidade": sub_quantity,
            "subordinadaPuResidual": totals["subordinada"],
            "deducoesTotal": 0,
            "provisoesTotal": totals["pdd"],
            "cotas": [
                {"classe": "SR1", "label": "Funding SR/MEZ", "tipo": "sr", "quantidade": 1, "taxa": "-", "pu": totals["funding"], "valor": totals["funding"]},
                {"classe": "SUB", "label": "Subordinadas", "tipo": "sub", "quantidade": 1, "taxa": "-", "pu": totals["subordinada"], "valor": totals["subordinada"]},
            ],
        },
        "carteiraResumo": {
            "valorNominal": totals["valorNominal"],
            "valorPresente": totals["carteiraVp"] + totals["pdd"],
            "valorPresenteLiquido": totals["carteiraVp"],
            "pddTotal": totals["pdd"],
            "cedentesUnicos": None,
            "sacadosUnicos": None,
            "prazoMedioDias": weighted_average(asset_rows, "prazoMedio", "carteiraVp"),
            "taxaMediaPonderada": weighted_average(asset_rows, "taxaMedia", "carteiraVp"),
            "preFixado": {"valorPresente": 0},
            "posFixado": {"valorPresente": 0},
        },
        "portfolioOverview": {
            "metrics": [
                {"label": "Carteira VP liquida", "value": currency(totals["carteiraVp"])},
                {"label": "Caixa total", "value": currency(totals["caixa"])},
                {"label": "Subordinadas", "value": currency(totals["subordinada"])},
                {"label": "PDD", "value": currency(totals["pdd"])},
            ],
            "assetRows": asset_rows,
            "rows": series_rows,
            "subHistory": sub_history,
        },
        "performanceCotas": [
            {"classe": "SR1", "label": "Funding SR/MEZ", "quantidade": 1, "taxa": "-", "pu": totals["funding"], "valor": totals["funding"], "resultadoDia": None},
            {"classe": "SUB", "label": "Subordinadas", "quantidade": 1, "taxa": "-", "pu": totals["subordinada"], "valor": totals["subordinada"], "resultadoDia": None},
        ],
        "rendimento30Dias": [],
        "sections": [{
            "id": "consolidado",
            "label": "Consolidado CRAs Carteira",
            "metrics": [
                {"label": "Ativo total", "value": currency(totals["ativoTotal"]), "isHighlight": True, "source": {"name": "Carteira + caixa"}},
                {"label": "Carteira VP liquida", "value": currency(totals["carteiraVp"]), "isHighlight": True, "source": {"name": "Import carteira"}},
                {"label": "Caixa total", "value": currency(totals["caixa"]), "isHighlight": True, "source": {"name": "Import caixa"}},
                {"label": "Funding SR/MEZ", "value": currency(totals["funding"]), "isHighlight": True, "source": {"name": "Memoria PU"}},
                {"label": "Subordinadas", "value": currency(totals["subordinada"]), "isHighlight": True, "source": {"name": "Residual"}},
                {"label": "PDD", "value": currency(totals["pdd"]), "isHighlight": True, "source": {"name": "Faixa vencimento"}},
            ],
        }],
        "sources": [
            {"name": "Importacao carteira+caixa CRAs Carteira", "date": report_date, "file": Path(source_meta["sourceFile"]).name},
        ],
    }
    return overview


def update_manifest(date_key: str, snapshots: dict[str, dict[str, Any]], overview: dict[str, Any]) -> None:
    manifest_path = REPO_ROOT / "data" / "cra-manifest.js"
    manifest = load_json_assignment(manifest_path)
    by_id = {item["craId"]: item for item in manifest}

    def make_date_entry(cra_id: str, key: str, snapshot: dict[str, Any]) -> dict[str, Any]:
        asset_imported = bool(snapshot.get("metadata", {}).get("assetImport"))
        return {
            "dateKey": key,
            "reportDate": snapshot.get("metadata", {}).get("reportDate"),
            "importedAt": snapshot.get("metadata", {}).get("importedAt"),
            "revisionId": snapshot.get("metadata", {}).get("revisionId"),
            "totalAtivo": snapshot.get("ativo", {}).get("total", 0),
            "funding": snapshot.get("passivo", {}).get("fundingTotal", 0),
            "subordinada": snapshot.get("passivo", {}).get("subordinadaTotal", 0),
            "dataScript": f"data/cras/{cra_id}/{key}.js",
            "assetImported": asset_imported,
            "portfolioAssetImport": asset_imported,
        }

    def sync_existing_dates(item: dict[str, Any], cra_id: str) -> None:
        folder = REPO_ROOT / "data" / "cras" / cra_id
        existing = {row.get("dateKey") for row in item.get("dates", [])}
        dates = list(item.get("dates", []))
        for path in sorted(folder.glob("*.js")):
            key = path.stem
            if not re.match(r"\d{4}-\d{2}-\d{2}$", key) or key in existing:
                continue
            try:
                snap = load_snapshot(path, cra_id, key)
            except Exception:
                continue
            dates.append(make_date_entry(cra_id, key, snap))
            existing.add(key)
        item["dates"] = sorted(dates, key=lambda row: row.get("dateKey", ""), reverse=True)

    def ensure_entry(cra_id: str, snapshot: dict[str, Any]) -> dict[str, Any]:
        item = by_id.get(cra_id)
        number = cra_number(cra_id)
        if item is None:
            item = {
                "craId": cra_id,
                "name": snapshot.get("cra", {}).get("name") or f"CRA Carteira {number}",
                "groupId": GROUP_ID,
                "groupName": GROUP_NAME,
                "childLabel": f"CRA {number}",
                "currentDate": date_key,
                "dates": [],
            }
            manifest.append(item)
            by_id[cra_id] = item
        item.setdefault("groupId", GROUP_ID)
        item.setdefault("groupName", GROUP_NAME)
        item.setdefault("childLabel", "Visao geral" if cra_id == "cras-carteira-overview" else f"CRA {number}")
        sync_existing_dates(item, cra_id)
        return item

    all_snapshots = {"cras-carteira-overview": overview, **snapshots}
    for cra_id, snapshot in all_snapshots.items():
        item = ensure_entry(cra_id, snapshot)
        item["currentDate"] = date_key
        date_entry = make_date_entry(cra_id, date_key, snapshot)
        dates = [row for row in item.get("dates", []) if row.get("dateKey") != date_key]
        dates.append(date_entry)
        item["dates"] = sorted(dates, key=lambda row: row.get("dateKey", ""), reverse=True)

    def manifest_key(item: dict[str, Any]) -> tuple[int, int]:
        if item.get("groupId") != GROUP_ID:
            return (0, 0)
        if item["craId"] == "cras-carteira-overview":
            return (1, 0)
        return (1, cra_number(item["craId"]) or 999)

    fixed = [item for item in manifest if item.get("groupId") != GROUP_ID]
    grouped = sorted([item for item in manifest if item.get("groupId") == GROUP_ID], key=manifest_key)
    write_manifest(manifest_path, fixed + grouped)


def cleanup_skipped_date(date_key: str, skipped_ids: set[str]) -> None:
    if not skipped_ids:
        return

    for cra_id in skipped_ids:
        path = REPO_ROOT / "data" / "cras" / cra_id / f"{date_key}.js"
        if path.exists():
            path.unlink()

    manifest_path = REPO_ROOT / "data" / "cra-manifest.js"
    manifest = load_json_assignment(manifest_path)
    for item in manifest:
        if item.get("craId") not in skipped_ids:
            continue
        dates = [row for row in item.get("dates", []) if row.get("dateKey") != date_key]
        item["dates"] = dates
        item["currentDate"] = dates[0]["dateKey"] if dates else ""
    write_manifest(manifest_path, manifest)


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa carteira+caixa consolidado dos CRAs Carteira.")
    parser.add_argument("source", nargs="?", default=str(DEFAULT_SOURCE), help="Arquivo .xlsx consolidado")
    args = parser.parse_args()
    source = Path(args.source)
    if not source.exists():
        raise FileNotFoundError(source)

    date_key, carteira_by_cra, cash_by_cra, source_meta = parse_workbook(source)
    snapshots: dict[str, dict[str, Any]] = {}
    active_ids = {
        cra_id
        for cra_id, rows in carteira_by_cra.items()
        if sum(to_number(row.get("valorPresenteDia")) for row in rows) > ACTIVE_VP_THRESHOLD
    }
    skipped_ids = (set(carteira_by_cra) | set(cash_by_cra)) - active_ids
    cleanup_skipped_date(date_key, skipped_ids)
    imported_ids = sorted(active_ids, key=lambda cid: cra_number(cid) or 0)
    for cra_id in imported_ids:
        rows = carteira_by_cra.get(cra_id, [])
        cash = cash_by_cra.get(cra_id, {"accounts": {}, "total": 0, "totalCalculado": 0, "rawRows": []})
        snapshot = build_snapshot(cra_id, date_key, rows, cash, source_meta)
        snapshots[cra_id] = snapshot
        write_snapshot(REPO_ROOT / "data" / "cras" / cra_id / f"{date_key}.js", cra_id, date_key, snapshot)

    overview = build_overview(date_key, snapshots, source_meta)
    write_snapshot(REPO_ROOT / "data" / "cras" / "cras-carteira-overview" / f"{date_key}.js", "cras-carteira-overview", date_key, overview)
    update_manifest(date_key, snapshots, overview)

    missing_carteira = sorted(set(cash_by_cra) - set(carteira_by_cra), key=lambda cid: cra_number(cid) or 0)
    missing_cash = sorted(set(carteira_by_cra) - set(cash_by_cra), key=lambda cid: cra_number(cid) or 0)
    print(json.dumps({
        "dateKey": date_key,
        "snapshots": len(snapshots),
        "overview": "cras-carteira-overview",
        "skippedSemCarteira": sorted(skipped_ids, key=lambda cid: cra_number(cid) or 0),
        "missingCarteira": missing_carteira,
        "missingCash": missing_cash,
        "source": str(source),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
